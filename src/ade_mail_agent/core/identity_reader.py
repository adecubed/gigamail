"""
identity_reader.py — Legge file e cartelle configurati nell'identity account.
Supporta: .txt .md .csv .xlsx .xls .pdf .docx
Usato da smart_draft per iniettare contesto nel prompt LLM.
"""
import os
import re
from typing import List, Dict

MAX_FILE_CHARS = 3000   # max caratteri per file singolo
MAX_TOTAL_CHARS = 12000  # max totale per tutti i file
MAX_FILES_PER_FOLDER = 20  # max file per cartella

SUPPORTED_EXTENSIONS = {'.txt', '.md', '.csv', '.xlsx', '.xls', '.pdf', '.docx'}


def read_identity_files(file_paths: List[str]) -> str:
    """
    Legge tutti i file e cartelle nella lista.
    Ritorna stringa formattata pronta per il prompt LLM.
    """
    if not file_paths:
        return ""

    results = []
    total_chars = 0

    for path in file_paths:
        if total_chars >= MAX_TOTAL_CHARS:
            break
        path = path.strip()
        if not path or not os.path.exists(path):
            continue

        if os.path.isdir(path):
            # Scansiona cartella
            folder_results = _read_folder(path, MAX_TOTAL_CHARS - total_chars)
            results.extend(folder_results)
            total_chars += sum(len(r['content']) for r in folder_results)
        else:
            # File singolo
            content = _read_file(path)
            if content:
                content = content[:MAX_FILE_CHARS]
                results.append({
                    'name': os.path.basename(path),
                    'path': path,
                    'type': 'file',
                    'content': content,
                })
                total_chars += len(content)

    if not results:
        return ""

    lines = ["FILE E CARTELLE DI RIFERIMENTO:"]
    for r in results:
        lines.append(f"\n--- {r['name']} ---")
        lines.append(r['content'])

    return "\n".join(lines)


def _read_folder(folder_path: str, max_chars: int) -> List[Dict]:
    """Scansiona cartella e legge i file supportati."""
    results = []
    total = 0

    try:
        entries = sorted(os.listdir(folder_path))
    except Exception:
        return []

    count = 0
    for name in entries:
        if count >= MAX_FILES_PER_FOLDER or total >= max_chars:
            break
        ext = os.path.splitext(name)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            continue
        full_path = os.path.join(folder_path, name)
        if not os.path.isfile(full_path):
            continue
        content = _read_file(full_path)
        if content:
            content = content[:MAX_FILE_CHARS]
            results.append({
                'name': f"{os.path.basename(folder_path)}/{name}",
                'path': full_path,
                'type': 'folder_file',
                'content': content,
            })
            total += len(content)
            count += 1

    return results


def _read_file(path: str) -> str:
    """Legge un file e ritorna il contenuto come testo."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in ('.txt', '.md'):
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read().strip()

        elif ext == '.csv':
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read().strip()

        elif ext in ('.xlsx', '.xls'):
            return _read_excel(path)

        elif ext == '.pdf':
            return _read_pdf(path)

        elif ext == '.docx':
            return _read_docx(path)

    except Exception as e:
        print(f"[IDENTITY READER] errore lettura {path}: {e}")
    return ""


def _read_excel(path: str) -> str:
    """Legge Excel e converte in testo tabellare."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames[:3]:  # max 3 fogli
            ws = wb[sheet]
            lines.append(f"[Foglio: {sheet}]")
            for row in ws.iter_rows(max_row=100, values_only=True):
                cells = [str(c) if c is not None else '' for c in row]
                line = ' | '.join(cells).strip(' |')
                if line:
                    lines.append(line)
        return '\n'.join(lines)
    except ImportError:
        # Fallback con xlrd per .xls
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            lines = []
            for sheet in wb.sheets()[:3]:
                lines.append(f"[Foglio: {sheet.name}]")
                for r in range(min(sheet.nrows, 100)):
                    cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    line = ' | '.join(cells).strip(' |')
                    if line:
                        lines.append(line)
            return '\n'.join(lines)
        except Exception as e:
            return f"[Excel non leggibile: {e}]"
    except Exception as e:
        return f"[Excel errore: {e}]"


def _read_pdf(path: str) -> str:
    """Legge PDF ed estrae testo."""
    try:
        import pypdf
        reader = pypdf.PdfReader(path)
        pages = []
        for page in reader.pages[:10]:  # max 10 pagine
            text = page.extract_text() or ''
            if text.strip():
                pages.append(text.strip())
        return '\n\n'.join(pages)
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                pages = []
                for page in pdf.pages[:10]:
                    text = page.extract_text() or ''
                    if text.strip():
                        pages.append(text.strip())
                return '\n\n'.join(pages)
        except Exception as e:
            return f"[PDF non leggibile: {e}]"
    except Exception as e:
        return f"[PDF errore: {e}]"


def _read_docx(path: str) -> str:
    """Legge DOCX ed estrae testo."""
    try:
        import docx
        doc = docx.Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return '\n'.join(paragraphs[:200])
    except Exception as e:
        return f"[DOCX errore: {e}]"


def list_folder_files(folder_path: str) -> List[Dict]:
    """
    Elenca i file in una cartella (per UI preview).
    Ritorna lista di {name, path, ext, size_kb}
    """
    if not os.path.isdir(folder_path):
        return []
    results = []
    try:
        for name in sorted(os.listdir(folder_path))[:50]:
            ext = os.path.splitext(name)[1].lower()
            if ext not in SUPPORTED_EXTENSIONS:
                continue
            full_path = os.path.join(folder_path, name)
            if os.path.isfile(full_path):
                size_kb = round(os.path.getsize(full_path) / 1024, 1)
                results.append({
                    'name': name,
                    'path': full_path,
                    'ext': ext,
                    'size_kb': size_kb,
                })
    except Exception:
        pass
    return results


def find_relevant_files(file_paths: List[str], query: str, max_files: int = 5) -> List[Dict]:
    """
    Cerca file rilevanti nelle cartelle configurate nell'identity.
    Matching per nome file vs parole chiave nella query.
    Usato da smart_draft per suggerire allegati automatici.
    
    Ritorna lista di {name, path, score} ordinata per rilevanza.
    """
    if not file_paths or not query:
        return []

    # Due famiglie di token:
    # - CODICI: sigle tipo "A.2.1", "B08", "a-12" — peso alto, confronto con
    #   separatori normalizzati (a.2.1 == a21 == a_2_1)
    # - PAROLE: >= 4 char, per evitare che "al"/"la" matchino dentro
    #   "balcone" o simili
    query_lower = query.lower()
    import re as _re
    codes = _re.findall(r'\b[a-z]{1,3}[\s._\-]?\d+(?:[._\-]\d+)*\b', query_lower)
    words = _re.findall(r'[a-z]{4,}', query_lower)
    tokens = {'codes': [_norm_code(c) for c in codes], 'words': words}
    if not tokens['codes'] and not tokens['words']:
        return []

    candidates = []

    for path in file_paths:
        path = path.strip()
        if not path or not os.path.exists(path):
            continue

        if os.path.isdir(path):
            # Scansiona cartella
            try:
                for name in os.listdir(path):
                    ext = os.path.splitext(name)[1].lower()
                    if ext not in SUPPORTED_EXTENSIONS and ext not in ('.jpg', '.jpeg', '.png'):
                        continue
                    full_path = os.path.join(path, name)
                    if not os.path.isfile(full_path):
                        continue
                    score = _score_file(name, tokens)
                    if score > 0:
                        candidates.append({
                            'name': name,
                            'path': full_path,
                            'score': score,
                            'ext': ext,
                        })
            except Exception:
                pass
        else:
            # File singolo
            name = os.path.basename(path)
            score = _score_file(name, tokens)
            if score > 0:
                candidates.append({
                    'name': name,
                    'path': path,
                    'score': score,
                    'ext': os.path.splitext(name)[1].lower(),
                })

    # Ordina per score decrescente
    candidates.sort(key=lambda x: x['score'], reverse=True)
    return candidates[:max_files]


def _norm_code(s: str) -> str:
    """Normalizza un codice rimuovendo i separatori: 'a.2.1' -> 'a21'."""
    import re as _re
    return _re.sub(r'[^a-z0-9]', '', s.lower())


def _score_file(filename: str, tokens) -> float:
    """
    Score di rilevanza nome-file vs token della query.
    tokens = {'codes': [codici normalizzati], 'words': [parole >=4 char]}
    """
    name_lower = os.path.splitext(filename.lower())[0]
    name_norm = _norm_code(name_lower)

    score = 0.0
    for code in tokens.get('codes', []):
        # confronto con separatori normalizzati: 'a21' trova 'A.2.1.pdf',
        # 'a_2_1.pdf', 'a-2-1 no balcone.pdf'
        if code and code in name_norm:
            score += 5.0
    for word in tokens.get('words', []):
        if word in name_lower:
            score += 1.0
    return score


def list_all_files(file_paths: List[str]) -> List[Dict]:
    """
    Lista tutti i file disponibili nelle cartelle configurate.
    Usato per dare all'LLM la lista completa da cui scegliere.
    """
    results = []
    for path in file_paths:
        path = path.strip()
        if not path or not os.path.exists(path):
            continue
        if os.path.isdir(path):
            try:
                for name in sorted(os.listdir(path)):
                    ext = os.path.splitext(name)[1].lower()
                    if ext not in SUPPORTED_EXTENSIONS and ext not in ('.jpg', '.jpeg', '.png'):
                        continue
                    full_path = os.path.join(path, name)
                    if os.path.isfile(full_path):
                        results.append({
                            'name': name,
                            'name_no_ext': os.path.splitext(name)[0],
                            'path': full_path,
                            'folder': os.path.basename(path),
                            'ext': ext,
                        })
            except Exception:
                pass
        elif os.path.isfile(path):
            ext = os.path.splitext(path)[1].lower()
            results.append({
                'name': os.path.basename(path),
                'name_no_ext': os.path.splitext(os.path.basename(path))[0],
                'path': path,
                'folder': '',
                'ext': ext,
            })
    return results


def find_files_by_names(file_paths: List[str], names: List[str]) -> List[Dict]:
    """
    Cerca file per lista di nomi suggeriti dall'LLM.
    Matching fuzzy: cerca il nome come sottostringa del nome file.
    """
    all_files = list_all_files(file_paths)
    results = []
    seen_paths = set()

    for name in names:
        name_clean = name.strip().lower()
        if not name_clean:
            continue
        # Rimuovi estensione se presente
        name_no_ext = os.path.splitext(name_clean)[0]

        for f in all_files:
            if f['path'] in seen_paths:
                continue
            fname_lower = f['name_no_ext'].lower()
            # Match: il nome cercato è contenuto nel nome file o viceversa
            if name_no_ext in fname_lower or fname_lower in name_no_ext:
                results.append(f)
                seen_paths.add(f['path'])

    return results
